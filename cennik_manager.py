"""
cennik_manager.py
=================
Modul pre správu cenníka aplikácie FINES Konfigurátor.

Funkcie:
  - export_do_master_excelu()  → prečíta všetkých 9 program_*.xlsx a vráti bytes
                                  jedného cennik_master.xlsx na stiahnutie.
  - import_z_master_excelu(uploaded_file) → overí a zapíše nový cenník
                                             zo súboru nahratého používateľom.

Štruktúra Master Excelu (9 záložiek):
  Sedacky | Postele_Manzelske | Postele_Sirky | Masiv |
  Pohovky | Jednolozka | Rosty | Doplnky | Stolicky
"""

import os
import io
import shutil
from datetime import datetime
import pandas as pd


# ---------------------------------------------------------------------------
# Definícia zdrojových súborov a ich záložiek v master exceli
# ---------------------------------------------------------------------------

SUBORY = {
    "Sedacky":           "program_all.xlsx",
    "Postele_Manzelske": "program_postele_v2.xlsx",
    "Postele_Sirky":     "program_postele_sirky.xlsx",
    "Masiv":             "program_masiv.xlsx",
    "Pohovky":           "program_pohovky.xlsx",
    "Jednolozka":        "program_jednolozka.xlsx",
    "Rosty":             "program_rosty.xlsx",
    "Doplnky":           "program_doplnky.xlsx",
    "Stolicky":          "program_stolicky.xlsx",
}

# Povinné stĺpce pre každú záložku – overujú sa pri importe
POVINNE_STLPCE = {
    "Sedacky":           ["Model", "Kod", "Popis", "Pohoda", "Zivot", "Neha"],
    "Postele_Manzelske": ["Kategoria", "Kod", "Popis", "Pohoda", "Zivot", "Neha"],
    "Postele_Sirky":     ["Model", "Kod", "Popis", "Pohoda", "Zivot", "Neha"],
    "Masiv":             ["Kod", "Popis", "Sirka", "Buk_Prirodny", "Buk_Moreny", "Dub_Farebny"],
    "Pohovky":           ["Podkategoria", "Model", "Kod", "Popis", "Pohoda", "Zivot", "Neha"],
    "Jednolozka":        ["Podkategoria", "Model", "Kod", "Popis", "Pohoda", "Zivot", "Neha"],
    "Rosty":             ["Kod", "Sirka", "Nosnost", "Cena_MOC", "Cena_MOC_Akcia"],
    "Doplnky":           ["Kategoria", "Kod", "Popis", "Pohoda", "Zivot", "Neha", "Cena_Fix"],
    "Stolicky":          ["Kod", "Popis", "Pohoda", "Zivot", "Neha", "Cena_Fix"],
}

# Numerické stĺpce – musia obsahovať čísla (nie text)
NUMERICKE_STLPCE = {
    "Sedacky":           ["Pohoda", "Zivot", "Neha"],
    "Postele_Manzelske": ["Pohoda", "Zivot", "Neha"],
    "Postele_Sirky":     ["Pohoda", "Zivot", "Neha"],
    "Masiv":             ["Buk_Prirodny", "Buk_Moreny", "Dub_Farebny"],
    "Pohovky":           ["Pohoda", "Zivot", "Neha"],
    "Jednolozka":        ["Pohoda", "Zivot", "Neha"],
    "Rosty":             ["Cena_MOC", "Cena_MOC_Akcia"],
    "Doplnky":           ["Pohoda", "Zivot", "Neha", "Cena_Fix"],
    "Stolicky":          ["Pohoda", "Zivot", "Neha", "Cena_Fix"],
}


# ---------------------------------------------------------------------------
# EXPORT
# ---------------------------------------------------------------------------

def export_do_master_excelu() -> bytes:
    """
    Prečíta všetkých 9 program_*.xlsx a zapíše ich do jedného cennik_master.xlsx.
    Vráti súbor ako bytes vhodné pre st.download_button.
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for zalozka, subor in SUBORY.items():
            if os.path.exists(subor):
                try:
                    df = pd.read_excel(subor)
                    # Zabezpečíme, že povinné stĺpce sú vždy prítomné
                    for stlpec in POVINNE_STLPCE.get(zalozka, []):
                        if stlpec not in df.columns:
                            df[stlpec] = ""
                    df.to_excel(writer, sheet_name=zalozka, index=False)
                except Exception as e:
                    pd.DataFrame(columns=POVINNE_STLPCE.get(zalozka, ["Chyba"])).to_excel(
                        writer, sheet_name=zalozka, index=False
                    )
            else:
                pd.DataFrame(columns=POVINNE_STLPCE.get(zalozka, [])).to_excel(
                    writer, sheet_name=zalozka, index=False
                )
    
    output.seek(0)
    return output.read()


def export_formatovany_cennik(platnost: str = "") -> bytes:
    """
    Vygeneruje profesionálny, formátovaný Excel cenník FINES s farebnou úpravou,
    logom (ak existuje), hlavičkou a všetkými kategóriami na oddielených záložkách.
    Tento súbor je určený na čítanie / tlač / zdieľanie, nie na import.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.worksheet.page import PageMargins
    import datetime

    # --- Farby ---
    COL_HLAVICKA_BG   = "1A3A6B"   # Tmavomodrá – hlavička tabuľky
    COL_HLAVICKA_TEXT = "FFFFFF"   # Biela
    COL_KATEGORIA_BG  = "2E6FBF"   # Stredná modrá – nadpis kategórie
    COL_KATEGORIA_TEXT = "FFFFFF"
    COL_ZEBRA_A       = "F0F4FA"   # Svetlá zebra
    COL_ZEBRA_B       = "FFFFFF"   # Biela zebra
    COL_TITULEK_BG    = "0D2240"   # Tmavá hlavička stránky
    COL_BORDER        = "BDD0EA"

    BORDER_THIN = Border(
        left=Side(style='thin', color=COL_BORDER),
        right=Side(style='thin', color=COL_BORDER),
        top=Side(style='thin', color=COL_BORDER),
        bottom=Side(style='thin', color=COL_BORDER),
    )

    if not platnost:
        platnost = datetime.date.today().strftime("%d.%m.%Y")

    wb = Workbook()
    wb.remove(wb.active)  # odstráni default záložku

    # Definícia záložiek a ich obsahu
    konfig = [
        {
            "zalozka_nazov": "Sedačky",
            "subor": "program_all.xlsx",
            "nadpis": "SEDACIE SÚPRAVY",
            "popis": "Ceny sú uvedené v € s DPH pre kategórie látok: Pohoda / Život / Neha",
            "skupiny": "Model",
            "stlpce": [
                ("Kod", "Kód", 12),
                ("Popis", "Popis elementu", 40),
                ("Pohoda", "Pohoda (€)", 13),
                ("Zivot", "Život (€)", 13),
                ("Neha", "Neha (€)", 13),
            ],
        },
        {
            "zalozka_nazov": "Manž. postele",
            "subor": "program_postele_v2.xlsx",
            "nadpis": "ČALÚNENÉ MANŽELSKÉ POSTELE (160/180 cm)",
            "popis": "Ceny zahŕňajú všetky varianty roštov a výklopov. Šírka 160 a 180 cm – rovnaká cena.",
            "skupiny": "Kategoria",
            "stlpce": [
                ("Kod", "Kód", 24),
                ("Popis", "Popis", 44),
                ("Pohoda", "Pohoda (€)", 13),
                ("Zivot", "Život (€)", 13),
                ("Neha", "Neha (€)", 13),
            ],
        },
        {
            "zalozka_nazov": "Postele 90–140",
            "subor": "program_postele_sirky.xlsx",
            "nadpis": "ČALÚNENÉ POSTELE (šírky 90–140 cm)",
            "popis": "Modely: BEATRIX, LIPARI, CREATIVE, NORIKA, DOROTEA + Denné deky",
            "skupiny": "Model",
            "stlpce": [
                ("Kod", "Kód", 24),
                ("Popis", "Popis", 42),
                ("Pohoda", "Pohoda (€)", 13),
                ("Zivot", "Život (€)", 13),
                ("Neha", "Neha (€)", 13),
            ],
        },
        {
            "zalozka_nazov": "Masívne postele",
            "subor": "program_masiv.xlsx",
            "nadpis": "DREVENÉ A MASÍVNE POSTELE",
            "popis": "Prevedenia dreva: Buk prírodný / Buk morený / Dub farebný",
            "skupiny": None,
            "stlpce": [
                ("Kod", "Kód", 20),
                ("Popis", "Popis", 36),
                ("Sirka", "Šírka", 10),
                ("Buk_Prirodny", "Buk prírodný (€)", 16),
                ("Buk_Moreny", "Buk morený (€)", 15),
                ("Dub_Farebny", "Dub farebný (€)", 15),
            ],
        },
        {
            "zalozka_nazov": "Pohovky a kreslá",
            "subor": "program_pohovky.xlsx",
            "nadpis": "ROZKLADACIE KRESLÁ A POHOVKY",
            "popis": "Modely: BART, TERI, MANIA, DARCY, GREMIA, FLAVIO, ARKÁDIA, POLINA, KORA, PRIMAVERA",
            "skupiny": "Podkategoria",
            "stlpce": [
                ("Model", "Model", 16),
                ("Kod", "Kód", 16),
                ("Popis", "Popis", 34),
                ("Pohoda", "Pohoda (€)", 13),
                ("Zivot", "Život (€)", 13),
                ("Neha", "Neha (€)", 13),
            ],
        },
        {
            "zalozka_nazov": "Jednolôžka",
            "subor": "program_jednolozka.xlsx",
            "nadpis": "JEDNOLÔŽKA A VÁĽANDY",
            "popis": "Modely: MIRAGE, OPTIMA, ROYA, SADIE, KAMI, NEO, IDA PLUS",
            "skupiny": "Podkategoria",
            "stlpce": [
                ("Model", "Model", 16),
                ("Kod", "Kód", 18),
                ("Popis", "Popis", 34),
                ("Pohoda", "Pohoda (€)", 13),
                ("Zivot", "Život (€)", 13),
                ("Neha", "Neha (€)", 13),
            ],
        },
        {
            "zalozka_nazov": "Rošty",
            "subor": "program_rosty.xlsx",
            "nadpis": "ROŠTY",
            "popis": "Veľkoobchodné ceny roštov podľa šírky a nosnosti",
            "skupiny": None,
            "stlpce": [
                ("Kod", "Kód", 16),
                ("Sirka", "Šírka", 12),
                ("Nosnost", "Nosnosť", 14),
                ("Cena_MOC", "Cena MOC (€)", 14),
                ("Cena_MOC_Akcia", "Cena MOC akcia (€)", 18),
            ],
        },
        {
            "zalozka_nazov": "Doplnky",
            "subor": "program_doplnky.xlsx",
            "nadpis": "DOPLNKY A SPÁLŇOVÝ NÁBYTOK",
            "popis": "Úložné priestory, nočné stolíky, denné deky a príslušenstvo",
            "skupiny": "Kategoria",
            "stlpce": [
                ("Kod", "Kód", 16),
                ("Popis", "Popis", 40),
                ("Pohoda", "Pohoda (€)", 13),
                ("Zivot", "Život (€)", 13),
                ("Neha", "Neha (€)", 13),
                ("Cena_Fix", "Fixná cena (€)", 14),
            ],
        },
        {
            "zalozka_nazov": "Stoličky a kreslá",
            "subor": "program_stolicky.xlsx",
            "nadpis": "JEDÁLENSKÉ STOLIČKY A RELAXAČNÉ KRESLÁ",
            "popis": "Stoličky a relaxačné kreslá z veľkoobchodného cenníka FINES",
            "skupiny": None,
            "stlpce": [
                ("Kod", "Kód", 14),
                ("Popis", "Popis", 44),
                ("Pohoda", "Pohoda (€)", 13),
                ("Zivot", "Život (€)", 13),
                ("Neha", "Neha (€)", 13),
                ("Cena_Fix", "Fixná cena (€)", 14),
            ],
        },
    ]

    for kat in konfig:
        ws = wb.create_sheet(title=kat["zalozka_nazov"])
        ws.sheet_view.showGridLines = False
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        ws.page_setup.fitToWidth = 1

        # --- LOGO + HLAVIČKA (riadky 1–5) ---
        logo_path = "logo.png"
        logo_vlozene = False
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 130
                img.height = 45
                ws.add_image(img, "A1")
                logo_vlozene = True
            except Exception:
                pass

        # Titul záložky (stĺpec B–koniec)
        posledny_stlpec = len(kat["stlpce"])
        posledny_pismeno = get_column_letter(posledny_stlpec)

        # Riadok 1: Titulek
        ws.merge_cells(f"B1:{posledny_pismeno}1")
        cell_tit = ws["B1"]
        cell_tit.value = f"FINES, a.s. – Cenník {platnost}"
        cell_tit.font = Font(name="Calibri", bold=True, size=11, color="666666")
        cell_tit.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[1].height = 18

        # Riadok 2: Nadpis kategórie
        ws.merge_cells(f"A2:{posledny_pismeno}2")
        cell_nad = ws["A2"]
        cell_nad.value = kat["nadpis"]
        cell_nad.font = Font(name="Calibri", bold=True, size=16, color=COL_HLAVICKA_TEXT)
        cell_nad.fill = PatternFill("solid", fgColor=COL_TITULEK_BG)
        cell_nad.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 30

        # Riadok 3: Popis
        ws.merge_cells(f"A3:{posledny_pismeno}3")
        cell_pop = ws["A3"]
        cell_pop.value = kat["popis"]
        cell_pop.font = Font(name="Calibri", italic=True, size=10, color="444444")
        cell_pop.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[3].height = 16

        # Riadok 4: prázdny oddeľovač
        ws.row_dimensions[4].height = 6

        # --- Načítanie dát ---
        if not os.path.exists(kat["subor"]):
            ws["A5"].value = "⚠️ Dáta nie sú dostupné."
            continue

        try:
            df = pd.read_excel(kat["subor"])
        except Exception:
            ws["A5"].value = "⚠️ Chyba pri čítaní dát."
            continue

        # --- Záhlavie tabuľky ---
        riadok = 5
        for col_idx, (_, label, _) in enumerate(kat["stlpce"], start=1):
            cell = ws.cell(row=riadok, column=col_idx, value=label)
            cell.font = Font(name="Calibri", bold=True, size=10, color=COL_HLAVICKA_TEXT)
            cell.fill = PatternFill("solid", fgColor=COL_HLAVICKA_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER_THIN
        ws.row_dimensions[riadok].height = 22
        riadok += 1

        # --- Dáta (s prípadným zoskupením) ---
        skupiny_stlpec = kat.get("skupiny")
        
        if skupiny_stlpec and skupiny_stlpec in df.columns:
            skupiny = df[skupiny_stlpec].unique()
            for skupina in skupiny:
                # Riadok skupiny / kategórie
                ws.merge_cells(start_row=riadok, start_column=1, end_row=riadok, end_column=posledny_stlpec)
                cell_sk = ws.cell(row=riadok, column=1, value=str(skupina))
                cell_sk.font = Font(name="Calibri", bold=True, size=11, color=COL_KATEGORIA_TEXT)
                cell_sk.fill = PatternFill("solid", fgColor=COL_KATEGORIA_BG)
                cell_sk.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                cell_sk.border = BORDER_THIN
                ws.row_dimensions[riadok].height = 20
                riadok += 1

                podmnozina = df[df[skupiny_stlpec] == skupina]
                for zebra_i, (_, row) in enumerate(podmnozina.iterrows()):
                    zebra_bg = COL_ZEBRA_A if zebra_i % 2 == 0 else COL_ZEBRA_B
                    for col_idx, (klic, _, _) in enumerate(kat["stlpce"], start=1):
                        hodnota = row.get(klic, "")
                        if pd.isna(hodnota):
                            hodnota = ""
                        cell = ws.cell(row=riadok, column=col_idx, value=hodnota)
                        cell.font = Font(name="Calibri", size=10)
                        cell.fill = PatternFill("solid", fgColor=zebra_bg)
                        cell.border = BORDER_THIN
                        # Zarovnanie a formát čísel
                        if isinstance(hodnota, float) and hodnota > 0:
                            cell.number_format = '#,##0.00 "€"'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    ws.row_dimensions[riadok].height = 16
                    riadok += 1
        else:
            # Bez skupín – plochý zoznam
            for zebra_i, (_, row) in enumerate(df.iterrows()):
                zebra_bg = COL_ZEBRA_A if zebra_i % 2 == 0 else COL_ZEBRA_B
                for col_idx, (klic, _, _) in enumerate(kat["stlpce"], start=1):
                    hodnota = row.get(klic, "")
                    if pd.isna(hodnota):
                        hodnota = ""
                    cell = ws.cell(row=riadok, column=col_idx, value=hodnota)
                    cell.font = Font(name="Calibri", size=10)
                    cell.fill = PatternFill("solid", fgColor=zebra_bg)
                    cell.border = BORDER_THIN
                    if isinstance(hodnota, float) and hodnota > 0:
                        cell.number_format = '#,##0.00 "€"'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                ws.row_dimensions[riadok].height = 16
                riadok += 1

        # --- Šírky stĺpcov ---
        for col_idx, (_, _, sirka) in enumerate(kat["stlpce"], start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = sirka

        # --- Pätička ---
        riadok += 1
        ws.merge_cells(start_row=riadok, start_column=1, end_row=riadok, end_column=posledny_stlpec)
        cell_pat = ws.cell(row=riadok, column=1,
                           value=f"FINES, a.s.  |  Cenník platný od: {platnost}  |  Všetky ceny sú uvedené s DPH")
        cell_pat.font = Font(name="Calibri", size=9, italic=True, color="888888")
        cell_pat.alignment = Alignment(horizontal="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()



# ---------------------------------------------------------------------------
# IMPORT
# ---------------------------------------------------------------------------

def import_z_master_excelu(uploaded_file) -> tuple:
    """
    Načíta nahraný cennik_master.xlsx, overí ho a zapíše dáta späť do
    príslušných program_*.xlsx súborov.

    Vracia: (uspech: bool, sprava: str)
      - uspech=True  → všetko OK, sprava obsahuje štatistiku
      - uspech=False → chyba, sprava obsahuje popis problému
    """
    # 1. Načítanie súboru
    try:
        xls = pd.ExcelFile(uploaded_file)
    except Exception as e:
        return False, f"❌ Nepodarilo sa otvoriť súbor: {e}"
    
    # 2. Overenie, že všetky záložky existujú
    chybajuce_zalozky = [z for z in SUBORY.keys() if z not in xls.sheet_names]
    if chybajuce_zalozky:
        return False, (
            f"❌ V nahratom súbore chýbajú tieto záložky: **{', '.join(chybajuce_zalozky)}**\n\n"
            f"Záložky musia byť pomenované presne takto: {', '.join(SUBORY.keys())}"
        )
    
    # 3. Načítanie a overenie každej záložky
    nactane_df = {}
    for zalozka in SUBORY.keys():
        try:
            df = pd.read_excel(xls, sheet_name=zalozka)
        except Exception as e:
            return False, f"❌ Chyba pri čítaní záložky '{zalozka}': {e}"
        
        # Overenie povinných stĺpcov
        povinne = POVINNE_STLPCE.get(zalozka, [])
        chybajuce_stlpce = [s for s in povinne if s not in df.columns]
        if chybajuce_stlpce:
            return False, (
                f"❌ V záložke **'{zalozka}'** chýbajú stĺpce: **{', '.join(chybajuce_stlpce)}**\n\n"
                f"Názvy stĺpcov nesmú byť zmenené. Záložka musí mať stĺpce: {', '.join(povinne)}"
            )
        
        # Overenie numerických stĺpcov
        numericke = NUMERICKE_STLPCE.get(zalozka, [])
        for stlpec in numericke:
            if stlpec in df.columns:
                # Skúsime konvertovať – ak zlyhá, oznámime chybu
                problematicke = df[stlpec].apply(lambda x: not _je_cislo(x))
                pocet_chyb = problematicke.sum()
                if pocet_chyb > 0:
                    priklady = df[problematicke][stlpec].head(3).tolist()
                    return False, (
                        f"❌ V záložke **'{zalozka}'**, stĺpci **'{stlpec}'** "
                        f"sa nachádzajú {pocet_chyb} nečíselné hodnoty.\n\n"
                        f"Príklady problematických hodnôt: {priklady}\n\n"
                        f"Cenové stĺpce musia obsahovať len čísla (napr. 299.5, nie 'neuvedené')."
                    )
                
                # Overenie záporných cien
                zap = df[stlpec].apply(lambda x: _je_cislo(x) and float(x) < 0).sum()
                if zap > 0:
                    return False, (
                        f"❌ V záložke **'{zalozka}'**, stĺpci **'{stlpec}'** "
                        f"sa nachádza {zap} záporná cena. Skontrolujte prosím hodnoty."
                    )
        
        nactane_df[zalozka] = df
    
    # 4. Záloha pred prepísaním
    _zaloha_pred_importom()
    
    # 5. Zápis do súborov
    statistika = []
    for zalozka, ciel_subor in SUBORY.items():
        df = nactane_df[zalozka]
        try:
            df.to_excel(ciel_subor, index=False)
            statistika.append(f"✅ **{zalozka}**: {len(df)} položiek")
        except Exception as e:
            return False, f"❌ Nepodarilo sa zapísať súbor '{ciel_subor}': {e}"
    
    sprava = "✅ **Cenník bol úspešne aktualizovaný!**\n\n" + "\n".join(statistika)
    celkom = sum(len(nactane_df[z]) for z in SUBORY)
    sprava += f"\n\n---\n🔢 **Celkovo: {celkom} produktov** vo všetkých kategóriách."
    
    return True, sprava


# ---------------------------------------------------------------------------
# ŠTATISTIKA
# ---------------------------------------------------------------------------

def get_statistiku() -> dict:
    """
    Vráti slovník {záložka: počet_riadkov} pre aktuálny stav program_*.xlsx.
    """
    stat = {}
    for zalozka, subor in SUBORY.items():
        if os.path.exists(subor):
            try:
                df = pd.read_excel(subor)
                stat[zalozka] = len(df)
            except Exception:
                stat[zalozka] = -1  # chyba čítania
        else:
            stat[zalozka] = 0  # súbor neexistuje
    return stat


# ---------------------------------------------------------------------------
# Pomocné funkcie
# ---------------------------------------------------------------------------

def _je_cislo(hodnota) -> bool:
    """Vráti True ak je hodnota číslo alebo prázdna (NaN)."""
    if pd.isna(hodnota):
        return True  # prázdna bunka je OK
    try:
        float(str(hodnota).replace(",", ".").strip())
        return True
    except (ValueError, TypeError):
        return False


def _zaloha_pred_importom():
    """Zálohuje všetky program_*.xlsx do priečinka archive/ pred importom."""
    casova_znacka = datetime.now().strftime("%Y-%m-%d_%H-%M")
    archive_dir = "archive"
    os.makedirs(archive_dir, exist_ok=True)
    
    for zalozka, subor in SUBORY.items():
        if os.path.exists(subor):
            try:
                nazov_zalohy = f"{archive_dir}/zaloha_{casova_znacka}_{subor}"
                shutil.copy2(subor, nazov_zalohy)
            except Exception:
                pass  # záloha nie je kritická – pokračujeme


def get_zoznam_zaloh() -> list:
    """
    Vráti zoznam zálohovacích priečinkov/súborov v archive/.
    Vracia zoznam názvov súborov, zoradených od najnovšej po najstaršiu.
    """
    archive_dir = "archive"
    if not os.path.exists(archive_dir):
        return []
    subory = [f for f in os.listdir(archive_dir) if f.startswith("zaloha_") and f.endswith(".xlsx")]
    return sorted(subory, reverse=True)
